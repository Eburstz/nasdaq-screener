#!/usr/bin/env python3
"""
NASDAQ Stock Screener - Enhanced Edition
Comprehensive technical analysis with scoring, risk assessment, and actionable insights.

New Features:
- Signal strength scoring (1-10)
- Additional indicators: MACD, Bollinger Bands, Volume, ADX
- Market cap classification
- Risk assessment
- Price momentum analysis
- Better Excel formatting with top picks
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import time
import json
import os
import sys
import warnings
warnings.filterwarnings('ignore')

# ─── CONFIG ───────────────────────────────────────────────────────────────────
LOOKBACK_DAYS = 400        # Enough history for MA150
CROSSOVER_WINDOW = 5       # Days to look back for recent crossovers
RSI_PERIOD = 14
RSI_OVERBOUGHT = 70
RSI_OVERSOLD = 30
DIVERGENCE_LOOKBACK = 30   # Bars to look back for divergence peaks/troughs
COMBO_WINDOW = 10          # Days window to match crossover + divergence
BATCH_SIZE = 50            # Tickers per yfinance batch download
DELAY_BETWEEN_BATCHES = 1  # Seconds between batches
ADX_PERIOD = 14            # Average Directional Index period
MACD_FAST = 12
MACD_SLOW = 26
MACD_SIGNAL = 9
VOLUME_LOOKBACK = 20       # Days for average volume calculation

# ─── SECTOR TICKER LISTS ─────────────────────────────────────────────────────
RENEWABLE_ENERGY_TICKERS = [
    "ENPH", "SEDG", "FSLR", "RUN", "CSIQ", "JKS", "ARRY", "MAXN",
    "NOVA", "SHLS", "STEM", "BE", "PLUG", "BLDP", "CLNE", "ORA",
    "CWEN", "AY", "HASI", "NEP", "BEPC", "DQ", "GPRE", "REX",
    "AMSC", "AMPS", "FLNC", "ENVX", "QS", "CHPT", "EVGO", "BLNK",
    "SPWR", "AZRE", "TPIC",
]

NUCLEAR_ENERGY_TICKERS = [
    "CEG", "CCJ", "LEU", "SMR", "OKLO", "NNE", "UEC", "UUUU",
    "DNN", "URG", "NXE", "EU", "BWXT", "GEV",
]

AI_TICKERS = [
    "NVDA", "AMD", "AVGO", "INTC", "MRVL", "ARM", "QCOM", "MU",
    "MSFT", "GOOG", "GOOGL", "META", "AMZN", "AAPL", "CRM", "NOW",
    "SNOW", "PLTR", "AI", "PATH", "DDOG", "CRWD", "ZS", "PANW",
    "NET", "S", "SPLK", "SOUN", "BBAI", "BIGC", "UPST", "IONQ",
    "RGTI", "QUBT", "SMCI", "DELL", "HPE", "ANET", "CDNS", "SNPS",
    "ADSK", "ANSS", "TER", "ONTO", "WOLF", "ACLS", "LSCC", "MCHP",
    "NXPI", "ON", "MPWR", "MBLY", "LAZR", "INVZ", "LIDR",
]


def get_nasdaq_tickers():
    """Fetch all NASDAQ-listed tickers from NASDAQ's FTP."""
    print("Fetching NASDAQ ticker list...")
    try:
        url = "https://raw.githubusercontent.com/rreichel3/US-Stock-Symbols/main/nasdaq/nasdaq_tickers.txt"
        tickers = pd.read_csv(url, header=None)[0].tolist()
        tickers = [t.strip() for t in tickers if t.strip().isalpha() and len(t.strip()) <= 5]
        print(f"  Found {len(tickers)} NASDAQ tickers")
        return tickers
    except Exception as e:
        print(f"  Primary source failed ({e}), trying backup...")
        try:
            url = "https://api.nasdaq.com/api/screener/stocks?tableType=traded&exchange=NASDAQ&download=true"
            import urllib.request
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode())
            tickers = [row['symbol'] for row in data['data']['rows']
                       if row['symbol'].isalpha() and len(row['symbol']) <= 5]
            print(f"  Found {len(tickers)} NASDAQ tickers (backup)")
            return tickers
        except Exception as e2:
            print(f"  Backup also failed ({e2}), using fallback NASDAQ-100 + sectors")
            return list(set(
                ["AAPL","MSFT","GOOG","GOOGL","AMZN","NVDA","META","TSLA","AVGO","PEP",
                 "COST","ADBE","CSCO","NFLX","CMCSA","AMD","INTC","INTU","TXN","QCOM",
                 "AMGN","AMAT","BKNG","ISRG","MDLZ","ADI","REGN","VRTX","LRCX","PANW",
                 "MU","SNPS","KLAC","CDNS","MELI","CRWD","ABNB","FTNT","DASH","MNST",
                 "ORLY","KDP","NXPI","MCHP","CTAS","KHC","DXCM","AEP","ON","ROST",
                 "CPRT","PAYX","FAST","ODFL","EXC","EA","VRSK","CTSH","XEL","GEHC",
                 "IDXX","ZS","ANSS","CDW","DDOG","TEAM","BKR","FANG","ILMN","WBD",
                 "MRNA","DLTR","BIIB","SIRI","LCID","RIVN","ZM","ROKU","SNAP","PINS",
                 "PLTR","COIN","HOOD","MARA","RIOT","SQ","PYPL","SOFI","AFRM","NU"]
                + RENEWABLE_ENERGY_TICKERS + NUCLEAR_ENERGY_TICKERS + AI_TICKERS
            ))


def compute_rsi(series, period=14):
    """Compute RSI for a price series."""
    delta = series.diff()
    gain = delta.where(delta > 0, 0.0)
    loss = -delta.where(delta < 0, 0.0)
    avg_gain = gain.ewm(alpha=1/period, min_periods=period).mean()
    avg_loss = loss.ewm(alpha=1/period, min_periods=period).mean()
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi


def compute_macd(series, fast=12, slow=26, signal=9):
    """Compute MACD, signal line, and histogram."""
    ema_fast = series.ewm(span=fast, adjust=False).mean()
    ema_slow = series.ewm(span=slow, adjust=False).mean()
    macd = ema_fast - ema_slow
    signal_line = macd.ewm(span=signal, adjust=False).mean()
    histogram = macd - signal_line
    return macd, signal_line, histogram


def compute_bollinger_bands(series, period=20, std_dev=2):
    """Compute Bollinger Bands."""
    sma = series.rolling(period).mean()
    std = series.rolling(period).std()
    upper = sma + (std * std_dev)
    lower = sma - (std * std_dev)
    return upper, sma, lower


def compute_adx(high, low, close, period=14):
    """Compute Average Directional Index (trend strength)."""
    # True Range
    tr1 = high - low
    tr2 = abs(high - close.shift())
    tr3 = abs(low - close.shift())
    tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
    atr = tr.ewm(alpha=1/period, min_periods=period).mean()

    # Directional Movement
    up_move = high - high.shift()
    down_move = low.shift() - low

    plus_dm = np.where((up_move > down_move) & (up_move > 0), up_move, 0)
    minus_dm = np.where((down_move > up_move) & (down_move > 0), down_move, 0)

    plus_dm = pd.Series(plus_dm, index=close.index).ewm(alpha=1/period, min_periods=period).mean()
    minus_dm = pd.Series(minus_dm, index=close.index).ewm(alpha=1/period, min_periods=period).mean()

    plus_di = 100 * (plus_dm / atr)
    minus_di = 100 * (minus_dm / atr)

    dx = 100 * abs(plus_di - minus_di) / (plus_di + minus_di)
    adx = dx.ewm(alpha=1/period, min_periods=period).mean()

    return adx


def get_market_cap_category(ticker_info):
    """Classify stock by market cap."""
    try:
        market_cap = ticker_info.get('marketCap', 0)
        if market_cap >= 200_000_000_000:
            return "Mega Cap (>$200B)"
        elif market_cap >= 10_000_000_000:
            return "Large Cap ($10B-$200B)"
        elif market_cap >= 2_000_000_000:
            return "Mid Cap ($2B-$10B)"
        elif market_cap >= 300_000_000:
            return "Small Cap ($300M-$2B)"
        elif market_cap > 0:
            return "Micro Cap (<$300M)"
        else:
            return "Unknown"
    except:
        return "Unknown"


def find_local_extrema(series, order=5):
    """Find local maxima and minima indices in a series."""
    maxima = []
    minima = []
    vals = series.values
    for i in range(order, len(vals) - order):
        if np.all(vals[i] >= vals[i-order:i]) and np.all(vals[i] >= vals[i+1:i+order+1]):
            maxima.append(i)
        if np.all(vals[i] <= vals[i-order:i]) and np.all(vals[i] <= vals[i+1:i+order+1]):
            minima.append(i)
    return maxima, minima


def detect_rsi_divergence(close, rsi, lookback=30):
    """Detect RSI divergences with strength score."""
    if len(close) < lookback + 10:
        return None

    recent = slice(-lookback, None)
    close_r = close.iloc[recent].reset_index(drop=True)
    rsi_r = rsi.iloc[recent].reset_index(drop=True)
    dates_r = close.index[recent]

    maxima, minima = find_local_extrema(close_r, order=3)

    # Bearish divergence
    if len(maxima) >= 2:
        i1, i2 = maxima[-2], maxima[-1]
        if (close_r.iloc[i2] > close_r.iloc[i1] and
            rsi_r.iloc[i2] < rsi_r.iloc[i1] and
            rsi_r.iloc[i2] > RSI_OVERBOUGHT):
            # Strength based on RSI divergence magnitude
            rsi_diff = abs(rsi_r.iloc[i1] - rsi_r.iloc[i2])
            strength = min(10, int(rsi_diff / 2))  # 2 RSI points = 1 strength
            return ("Bearish Divergence", dates_r[i2], strength)

    # Bullish divergence
    if len(minima) >= 2:
        i1, i2 = minima[-2], minima[-1]
        if (close_r.iloc[i2] < close_r.iloc[i1] and
            rsi_r.iloc[i2] > rsi_r.iloc[i1] and
            rsi_r.iloc[i2] < RSI_OVERSOLD):
            rsi_diff = abs(rsi_r.iloc[i2] - rsi_r.iloc[i1])
            strength = min(10, int(rsi_diff / 2))
            return ("Bullish Divergence", dates_r[i2], strength)

    return None


def detect_ma_crossover(ma_short, ma_long, dates, window=5):
    """Detect MA crossover with strength score."""
    if len(ma_short) < window + 2:
        return None

    for i in range(-window, 0):
        prev_diff = ma_short.iloc[i-1] - ma_long.iloc[i-1]
        curr_diff = ma_short.iloc[i] - ma_long.iloc[i]
        if np.isnan(prev_diff) or np.isnan(curr_diff):
            continue

        if prev_diff <= 0 and curr_diff > 0:
            # Bullish cross strength based on separation
            separation_pct = (curr_diff / ma_long.iloc[i]) * 100
            strength = min(10, max(1, int(separation_pct * 20)))
            return ("Bullish Cross", dates[i], strength)

        if prev_diff >= 0 and curr_diff < 0:
            # Bearish cross
            separation_pct = abs(curr_diff / ma_long.iloc[i]) * 100
            strength = min(10, max(1, int(separation_pct * 20)))
            return ("Bearish Cross", dates[i], strength)

    return None


def calculate_signal_score(result):
    """
    Calculate overall signal score (0-10) based on multiple factors.
    Higher score = stronger signal quality.
    """
    score = 0
    factors = []

    # MA crossover signals (0-3 points each)
    if result.get('ma50_150_signal'):
        score += result.get('ma50_150_strength', 0) * 0.3
        factors.append(f"MA50×150 ({result.get('ma50_150_strength', 0)}/10)")

    if result.get('ma20_50_signal'):
        score += result.get('ma20_50_strength', 0) * 0.3
        factors.append(f"MA20×50 ({result.get('ma20_50_strength', 0)}/10)")

    # RSI divergence (0-3 points)
    if result.get('rsi_divergence'):
        score += result.get('rsi_div_strength', 0) * 0.3
        factors.append(f"RSI Div ({result.get('rsi_div_strength', 0)}/10)")

    # MACD confirmation (0-2 points)
    macd_signal = result.get('macd_signal')
    if macd_signal and 'Bullish' in str(result.get('ma50_150_signal', '')):
        score += 2
        factors.append("MACD Confirm")
    elif macd_signal and 'Bearish' in str(result.get('ma50_150_signal', '')):
        score += 2
        factors.append("MACD Confirm")

    # Volume confirmation (0-1.5 points)
    vol_ratio = result.get('volume_ratio', 1.0)
    if vol_ratio > 1.5:
        score += 1.5
        factors.append(f"Volume {vol_ratio:.1f}x")
    elif vol_ratio > 1.2:
        score += 1
        factors.append(f"Volume {vol_ratio:.1f}x")

    # ADX trend strength (0-1.5 points)
    adx = result.get('adx')
    if adx and adx > 40:
        score += 1.5
        factors.append(f"ADX {adx:.0f}")
    elif adx and adx > 25:
        score += 1
        factors.append(f"ADX {adx:.0f}")

    # Bollinger Band position (0-1 point)
    bb_position = result.get('bb_position')
    if bb_position and bb_position < 10:
        score += 1
        factors.append("Near Lower BB")
    elif bb_position and bb_position > 90:
        score += 1
        factors.append("Near Upper BB")

    return round(min(10, score), 1), factors


def assess_risk(result):
    """Assess risk level based on volatility and position."""
    risk_score = 0

    # Volatility (Bollinger Band width)
    bb_width = result.get('bb_width_pct', 0)
    if bb_width > 30:
        risk_score += 3  # High volatility = higher risk
    elif bb_width > 15:
        risk_score += 2
    else:
        risk_score += 1

    # RSI extremes
    rsi = result.get('last_rsi', 50)
    if rsi > 80 or rsi < 20:
        risk_score += 2
    elif rsi > 70 or rsi < 30:
        risk_score += 1

    # Market cap (smaller = higher risk)
    market_cap = result.get('market_cap_category', '')
    if 'Micro' in market_cap:
        risk_score += 3
    elif 'Small' in market_cap:
        risk_score += 2
    elif 'Mid' in market_cap:
        risk_score += 1

    # Momentum volatility
    momentum_5d = result.get('momentum_5d', 0)
    if abs(momentum_5d) > 15:
        risk_score += 1

    # Risk categorization
    if risk_score >= 7:
        return "High Risk"
    elif risk_score >= 4:
        return "Medium Risk"
    else:
        return "Low Risk"


def analyze_ticker(ticker, hist, ticker_info=None):
    """Enhanced ticker analysis with additional indicators."""
    if hist is None or len(hist) < 200:
        return None

    close = hist['Close']
    high = hist['High']
    low = hist['Low']
    volume = hist['Volume']
    dates = hist.index

    # Calculate indicators
    ma20 = close.rolling(20).mean()
    ma50 = close.rolling(50).mean()
    ma150 = close.rolling(150).mean()
    rsi = compute_rsi(close, RSI_PERIOD)
    macd, macd_signal, macd_hist = compute_macd(close, MACD_FAST, MACD_SLOW, MACD_SIGNAL)
    bb_upper, bb_middle, bb_lower = compute_bollinger_bands(close, 20, 2)
    adx = compute_adx(high, low, close, ADX_PERIOD)

    # Volume analysis
    avg_volume = volume.rolling(VOLUME_LOOKBACK).mean()
    volume_ratio = volume.iloc[-1] / avg_volume.iloc[-1] if avg_volume.iloc[-1] > 0 else 1.0

    # Price momentum
    momentum_5d = ((close.iloc[-1] / close.iloc[-6]) - 1) * 100 if len(close) >= 6 else 0
    momentum_20d = ((close.iloc[-1] / close.iloc[-21]) - 1) * 100 if len(close) >= 21 else 0
    momentum_60d = ((close.iloc[-1] / close.iloc[-61]) - 1) * 100 if len(close) >= 61 else 0

    # Bollinger Band position (0-100%)
    bb_width = bb_upper.iloc[-1] - bb_lower.iloc[-1]
    bb_position = ((close.iloc[-1] - bb_lower.iloc[-1]) / bb_width * 100) if bb_width > 0 else 50
    bb_width_pct = (bb_width / close.iloc[-1]) * 100  # Volatility measure

    # Market cap
    market_cap_category = get_market_cap_category(ticker_info) if ticker_info else "Unknown"

    result = {
        'ticker': ticker,
        'last_close': round(close.iloc[-1], 2),
        'last_rsi': round(rsi.iloc[-1], 2) if not np.isnan(rsi.iloc[-1]) else None,
        'ma20': round(ma20.iloc[-1], 2) if not np.isnan(ma20.iloc[-1]) else None,
        'ma50': round(ma50.iloc[-1], 2) if not np.isnan(ma50.iloc[-1]) else None,
        'ma150': round(ma150.iloc[-1], 2) if not np.isnan(ma150.iloc[-1]) else None,
        'macd': round(macd.iloc[-1], 2) if not np.isnan(macd.iloc[-1]) else None,
        'macd_signal_line': round(macd_signal.iloc[-1], 2) if not np.isnan(macd_signal.iloc[-1]) else None,
        'macd_histogram': round(macd_hist.iloc[-1], 2) if not np.isnan(macd_hist.iloc[-1]) else None,
        'adx': round(adx.iloc[-1], 2) if not np.isnan(adx.iloc[-1]) else None,
        'bb_upper': round(bb_upper.iloc[-1], 2),
        'bb_lower': round(bb_lower.iloc[-1], 2),
        'bb_position': round(bb_position, 1),
        'bb_width_pct': round(bb_width_pct, 2),
        'volume_ratio': round(volume_ratio, 2),
        'momentum_5d': round(momentum_5d, 2),
        'momentum_20d': round(momentum_20d, 2),
        'momentum_60d': round(momentum_60d, 2),
        'market_cap_category': market_cap_category,
    }

    # MACD signal
    if macd_hist.iloc[-1] > 0 and macd_hist.iloc[-2] <= 0:
        result['macd_signal'] = "Bullish MACD Cross"
    elif macd_hist.iloc[-1] < 0 and macd_hist.iloc[-2] >= 0:
        result['macd_signal'] = "Bearish MACD Cross"
    else:
        result['macd_signal'] = None

    # MA crossover signals with strength
    cross_50_150 = detect_ma_crossover(ma50, ma150, dates, CROSSOVER_WINDOW)
    if cross_50_150:
        result['ma50_150_signal'] = cross_50_150[0]
        result['ma50_150_date'] = cross_50_150[1].strftime('%Y-%m-%d')
        result['ma50_150_strength'] = cross_50_150[2]
    else:
        result['ma50_150_signal'] = None
        result['ma50_150_date'] = None
        result['ma50_150_strength'] = 0

    cross_20_50 = detect_ma_crossover(ma20, ma50, dates, CROSSOVER_WINDOW)
    if cross_20_50:
        result['ma20_50_signal'] = cross_20_50[0]
        result['ma20_50_date'] = cross_20_50[1].strftime('%Y-%m-%d')
        result['ma20_50_strength'] = cross_20_50[2]
    else:
        result['ma20_50_signal'] = None
        result['ma20_50_date'] = None
        result['ma20_50_strength'] = 0

    # RSI divergence with strength
    div = detect_rsi_divergence(close, rsi, DIVERGENCE_LOOKBACK)
    if div:
        result['rsi_divergence'] = div[0]
        result['rsi_div_date'] = div[1].strftime('%Y-%m-%d')
        result['rsi_div_strength'] = div[2]
    else:
        result['rsi_divergence'] = None
        result['rsi_div_date'] = None
        result['rsi_div_strength'] = 0

    # Combined signal
    result['combined_signal'] = None
    if cross_50_150 and div:
        days_apart = abs((cross_50_150[1] - div[1]).days)
        if days_apart <= COMBO_WINDOW:
            result['combined_signal'] = f"{cross_50_150[0]} + {div[0]}"

    # Calculate overall score and risk
    score, factors = calculate_signal_score(result)
    result['signal_score'] = score
    result['score_factors'] = ' | '.join(factors) if factors else 'No signals'
    result['risk_level'] = assess_risk(result)

    return result


def download_and_analyze(tickers, label="All NASDAQ"):
    """Download data in batches and analyze each ticker."""
    results = []
    total = len(tickers)
    start_date = (datetime.now() - timedelta(days=LOOKBACK_DAYS)).strftime('%Y-%m-%d')
    end_date = datetime.now().strftime('%Y-%m-%d')

    print(f"\nScanning {total} tickers for [{label}]...")

    for i in range(0, total, BATCH_SIZE):
        batch = tickers[i:i+BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE
        print(f"  Batch {batch_num}/{total_batches} ({len(batch)} tickers)...", end=" ", flush=True)

        try:
            data = yf.download(
                batch, start=start_date, end=end_date,
                group_by='ticker', progress=False, threads=True
            )
        except Exception as e:
            print(f"Error: {e}")
            time.sleep(DELAY_BETWEEN_BATCHES)
            continue

        for ticker in batch:
            try:
                if len(batch) == 1:
                    hist = data
                else:
                    hist = data[ticker].dropna()

                if hist is not None and len(hist) > 0:
                    # Try to get additional info
                    ticker_info = None
                    try:
                        ticker_obj = yf.Ticker(ticker)
                        ticker_info = ticker_obj.info
                    except:
                        pass

                    result = analyze_ticker(ticker, hist, ticker_info)
                    if result:
                        results.append(result)
            except Exception:
                pass

        print(f"done ({len(results)} analyzed so far)")
        if i + BATCH_SIZE < total:
            time.sleep(DELAY_BETWEEN_BATCHES)

    return results


def filter_signals(results, signal_type):
    """Filter results for a specific signal type."""
    if signal_type == 'ma50_150':
        return [r for r in results if r.get('ma50_150_signal')]
    elif signal_type == 'ma20_50':
        return [r for r in results if r.get('ma20_50_signal')]
    elif signal_type == 'rsi_div':
        return [r for r in results if r.get('rsi_divergence')]
    elif signal_type == 'combined':
        return [r for r in results if r.get('combined_signal')]
    elif signal_type == 'high_score':
        return [r for r in results if r.get('signal_score', 0) >= 6.0]
    return results


def apply_score_formatting(ws, row, score_col):
    """Apply conditional formatting based on signal score."""
    cell = ws.cell(row=row, column=score_col)
    score = cell.value

    if isinstance(score, (int, float)):
        if score >= 7.5:
            cell.fill = PatternFill('solid', fgColor='00B050')  # Dark green
            cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        elif score >= 6.0:
            cell.fill = PatternFill('solid', fgColor='92D050')  # Light green
            cell.font = Font(bold=True, color='000000', name='Arial', size=10)
        elif score >= 4.0:
            cell.fill = PatternFill('solid', fgColor='FFC000')  # Orange
            cell.font = Font(name='Arial', size=10)
        else:
            cell.fill = PatternFill('solid', fgColor='F0F0F0')  # Gray
            cell.font = Font(name='Arial', size=10)


def apply_risk_formatting(ws, row, risk_col):
    """Apply conditional formatting based on risk level."""
    cell = ws.cell(row=row, column=risk_col)
    risk = cell.value

    if risk == "Low Risk":
        cell.fill = PatternFill('solid', fgColor='C6EFCE')
        cell.font = Font(color='006100', name='Arial', size=10)
    elif risk == "Medium Risk":
        cell.fill = PatternFill('solid', fgColor='FFEB9C')
        cell.font = Font(color='9C5700', name='Arial', size=10)
    elif risk == "High Risk":
        cell.fill = PatternFill('solid', fgColor='FFC7CE')
        cell.font = Font(color='9C0006', name='Arial', size=10)


def style_header(ws, row, max_col):
    """Apply header styling."""
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    thin_border = Border(bottom=Side(style='thin', color='1F4E79'))

    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def write_enhanced_signal_sheet(wb, sheet_name, data, columns, col_widths):
    """Write enhanced signal sheet with score and risk columns."""
    ws = wb.create_sheet(title=sheet_name)

    # Headers
    for ci, col in enumerate(columns, 1):
        ws.cell(row=1, column=ci, value=col)
    style_header(ws, 1, len(columns))

    # Column widths
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data
    bullish_font = Font(color='006100', name='Arial', size=10)
    bearish_font = Font(color='9C0006', name='Arial', size=10)
    normal_font = Font(name='Arial', size=10)
    alt_fill = PatternFill('solid', fgColor='F2F7FB')

    # Find score and risk columns
    score_col = columns.index('Score') + 1 if 'Score' in columns else None
    risk_col = columns.index('Risk') + 1 if 'Risk' in columns else None

    for ri, row_data in enumerate(data, 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center')

            # Alternate row coloring (unless score/risk formatting applied)
            if ri % 2 == 0 and ci != score_col and ci != risk_col:
                cell.fill = alt_fill

            # Color signal cells
            if isinstance(val, str):
                if 'Bullish' in val:
                    cell.font = bullish_font
                elif 'Bearish' in val:
                    cell.font = bearish_font

        # Apply score and risk formatting
        if score_col:
            apply_score_formatting(ws, ri, score_col)
        if risk_col:
            apply_risk_formatting(ws, ri, risk_col)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data)+1}"
    ws.freeze_panes = 'A2'
    return ws


def build_excel(all_results, renewable_results, nuclear_results, ai_results, output_path):
    """Build enhanced Excel workbook with top picks and detailed analysis."""
    wb = Workbook()
    wb.remove(wb.active)

    scan_date = datetime.now().strftime('%Y-%m-%d %H:%M')

    # ─── TOP PICKS SHEET ──────────────────────────────────────────────────
    ws_picks = wb.create_sheet("🌟 TOP PICKS")

    # Title
    ws_picks['A1'] = "TOP PICKS - Highest Scoring Stocks"
    ws_picks['A1'].font = Font(bold=True, name='Arial', size=14, color='1F4E79')
    ws_picks['A2'] = f"Scan Date: {scan_date}"
    ws_picks['A2'].font = Font(name='Arial', size=10)

    # Get top scoring stocks (score >= 6.0)
    top_picks = sorted([r for r in all_results if r.get('signal_score', 0) >= 6.0],
                      key=lambda x: x.get('signal_score', 0), reverse=True)[:50]

    columns_picks = ["Rank", "Ticker", "Score", "Risk", "Price", "Signal", "Date",
                     "RSI", "MACD", "Vol", "5D%", "20D%", "Cap", "Key Factors"]
    widths_picks = [6, 8, 7, 12, 10, 20, 12, 7, 10, 6, 8, 8, 16, 50]

    for ci, col in enumerate(columns_picks, 1):
        ws_picks.cell(row=4, column=ci, value=col)
    style_header(ws_picks, 4, len(columns_picks))

    for ci, w in enumerate(widths_picks, 1):
        ws_picks.column_dimensions[get_column_letter(ci)].width = w

    for rank, stock in enumerate(top_picks, 1):
        row_data = [
            rank,
            stock['ticker'],
            stock['signal_score'],
            stock['risk_level'],
            stock['last_close'],
            stock.get('ma50_150_signal') or stock.get('ma20_50_signal') or stock.get('rsi_divergence', ''),
            stock.get('ma50_150_date') or stock.get('ma20_50_date') or stock.get('rsi_div_date', ''),
            stock['last_rsi'],
            stock.get('macd_signal', ''),
            stock['volume_ratio'],
            stock['momentum_5d'],
            stock['momentum_20d'],
            stock['market_cap_category'].split('(')[0].strip(),
            stock['score_factors']
        ]

        ri = rank + 4
        for ci, val in enumerate(row_data, 1):
            cell = ws_picks.cell(row=ri, column=ci, value=val)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='center' if ci < 13 else 'left')

            # Alternate row coloring
            if ri % 2 == 1:
                cell.fill = PatternFill('solid', fgColor='F2F7FB')

        # Apply score and risk formatting
        apply_score_formatting(ws_picks, ri, 3)
        apply_risk_formatting(ws_picks, ri, 4)

    ws_picks.freeze_panes = 'A5'
    ws_picks.auto_filter.ref = f"A4:{get_column_letter(len(columns_picks))}{len(top_picks)+4}"

    # ─── SUMMARY SHEET ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("📊 Summary")
    ws_sum['A1'] = "NASDAQ Stock Screener - Enhanced Analysis"
    ws_sum['A1'].font = Font(bold=True, name='Arial', size=14, color='1F4E79')
    ws_sum['A3'] = f"Scan Date: {scan_date}"
    ws_sum['A3'].font = Font(name='Arial', size=10)

    ws_sum['A5'] = "Quick Stats"
    ws_sum['A5'].font = Font(bold=True, name='Arial', size=12, color='1F4E79')

    datasets = [
        ("All NASDAQ", all_results),
        ("Renewable Energy", renewable_results),
        ("Nuclear Energy", nuclear_results),
        ("AI Companies", ai_results),
    ]

    row = 7
    for label, results in datasets:
        ma50_150 = filter_signals(results, 'ma50_150')
        ma20_50 = filter_signals(results, 'ma20_50')
        rsi_div = filter_signals(results, 'rsi_div')
        combined = filter_signals(results, 'combined')
        high_score = filter_signals(results, 'high_score')

        ws_sum.cell(row=row, column=1, value=label).font = Font(bold=True, name='Arial', size=10)
        ws_sum.cell(row=row, column=2, value=f"Tickers scanned: {len(results)}").font = Font(name='Arial', size=10)
        ws_sum.cell(row=row+1, column=2, value=f"High Score (≥6.0): {len(high_score)}").font = Font(name='Arial', size=10, color='00B050', bold=True)
        ws_sum.cell(row=row+2, column=2, value=f"MA50/MA150 crossovers: {len(ma50_150)}").font = Font(name='Arial', size=10)
        ws_sum.cell(row=row+3, column=2, value=f"MA20/MA50 crossovers: {len(ma20_50)}").font = Font(name='Arial', size=10)
        ws_sum.cell(row=row+4, column=2, value=f"RSI divergences: {len(rsi_div)}").font = Font(name='Arial', size=10)
        ws_sum.cell(row=row+5, column=2, value=f"Combined signals: {len(combined)}").font = Font(name='Arial', size=10)
        row += 7

    ws_sum['A30'] = "📖 How to Use This Report"
    ws_sum['A30'].font = Font(bold=True, name='Arial', size=12, color='1F4E79')

    guide = [
        "",
        "1. START HERE: Check the 'TOP PICKS' sheet for highest-scoring opportunities",
        "2. Score 7.5+: Strong signals with multiple confirmations - review first",
        "3. Score 6.0-7.5: Good signals worth investigating further",
        "4. Risk Level: Consider your risk tolerance (Low/Medium/High)",
        "5. Key Factors: Shows what contributed to the score",
        "",
        "Signal Types:",
        "• MA50×150: Longer-term trend changes (weeks to months)",
        "• MA20×50: Shorter-term momentum shifts (days to weeks)",
        "• RSI Divergence: Price/momentum conflicts (reversal warning)",
        "• Combined: Multiple signals together (strongest setups)",
        "",
        "Additional Indicators:",
        "• MACD: Trend momentum confirmation",
        "• Vol: Volume spike relative to average (>1.5x is significant)",
        "• ADX: Trend strength (>25 = trending, >40 = strong trend)",
        "• BB Position: Bollinger Band % (near 0 = oversold, near 100 = overbought)",
    ]

    for i, line in enumerate(guide, 31):
        ws_sum.cell(row=i, column=1, value=line).font = Font(name='Arial', size=10)

    ws_sum.column_dimensions['A'].width = 70
    ws_sum.column_dimensions['B'].width = 40

    # ─── DETAILED SIGNAL SHEETS ───────────────────────────────────────────
    columns_ma_enh = ["Ticker", "Score", "Risk", "Price", "MA Short", "MA Long",
                      "Signal", "Date", "RSI", "MACD", "Vol", "ADX", "5D%", "20D%"]
    widths_ma_enh = [8, 7, 12, 10, 10, 10, 18, 12, 7, 14, 6, 7, 8, 8]

    columns_rsi_enh = ["Ticker", "Score", "Risk", "Price", "RSI", "Divergence", "Date",
                       "MACD", "Vol", "ADX", "BB%", "5D%", "20D%"]
    widths_rsi_enh = [8, 7, 12, 10, 7, 22, 12, 14, 6, 7, 7, 8, 8]

    columns_combo_enh = ["Ticker", "Score", "Risk", "Price", "MA Signal", "MA Date",
                         "RSI Signal", "RSI Date", "Combined", "Vol", "ADX", "5D%", "Cap"]
    widths_combo_enh = [8, 7, 12, 10, 18, 12, 22, 12, 30, 6, 7, 8, 16]

    for label, results, prefix in [
        ("All NASDAQ", all_results, "All"),
        ("Renewable", renewable_results, "RE"),
        ("Nuclear", nuclear_results, "Nuc"),
        ("AI", ai_results, "AI"),
    ]:
        # High Score Sheet (NEW)
        high_score = sorted(filter_signals(results, 'high_score'),
                          key=lambda x: x.get('signal_score', 0), reverse=True)
        rows = [[r['ticker'], r['signal_score'], r['risk_level'], r['last_close'],
                 r.get('ma50_150_signal') or r.get('ma20_50_signal') or '',
                 r.get('ma50_150_date') or r.get('ma20_50_date') or '',
                 r.get('rsi_divergence', ''), r.get('rsi_div_date', ''),
                 r.get('combined_signal', ''),
                 r['volume_ratio'], r['adx'], r['momentum_5d'],
                 r['market_cap_category'].split('(')[0].strip()]
                for r in high_score]
        write_enhanced_signal_sheet(wb, f"{prefix} High Score", rows, columns_combo_enh, widths_combo_enh)

        # MA50 x MA150
        ma50_150 = sorted(filter_signals(results, 'ma50_150'),
                         key=lambda x: x.get('signal_score', 0), reverse=True)
        rows = [[r['ticker'], r['signal_score'], r['risk_level'], r['last_close'],
                 r['ma50'], r['ma150'], r['ma50_150_signal'], r['ma50_150_date'],
                 r['last_rsi'], r.get('macd_signal', ''), r['volume_ratio'],
                 r['adx'], r['momentum_5d'], r['momentum_20d']]
                for r in ma50_150]
        write_enhanced_signal_sheet(wb, f"{prefix} MA50×150", rows, columns_ma_enh, widths_ma_enh)

        # MA20 x MA50
        ma20_50 = sorted(filter_signals(results, 'ma20_50'),
                        key=lambda x: x.get('signal_score', 0), reverse=True)
        rows = [[r['ticker'], r['signal_score'], r['risk_level'], r['last_close'],
                 r['ma20'], r['ma50'], r['ma20_50_signal'], r['ma20_50_date'],
                 r['last_rsi'], r.get('macd_signal', ''), r['volume_ratio'],
                 r['adx'], r['momentum_5d'], r['momentum_20d']]
                for r in ma20_50]
        write_enhanced_signal_sheet(wb, f"{prefix} MA20×50", rows, columns_ma_enh, widths_ma_enh)

        # RSI Divergence
        rsi_div = sorted(filter_signals(results, 'rsi_div'),
                        key=lambda x: x.get('signal_score', 0), reverse=True)
        rows = [[r['ticker'], r['signal_score'], r['risk_level'], r['last_close'],
                 r['last_rsi'], r['rsi_divergence'], r['rsi_div_date'],
                 r.get('macd_signal', ''), r['volume_ratio'], r['adx'],
                 r['bb_position'], r['momentum_5d'], r['momentum_20d']]
                for r in rsi_div]
        write_enhanced_signal_sheet(wb, f"{prefix} RSI Div", rows, columns_rsi_enh, widths_rsi_enh)

        # Combined
        combined = sorted(filter_signals(results, 'combined'),
                         key=lambda x: x.get('signal_score', 0), reverse=True)
        rows = [[r['ticker'], r['signal_score'], r['risk_level'], r['last_close'],
                 r['ma50_150_signal'], r['ma50_150_date'],
                 r['rsi_divergence'], r['rsi_div_date'],
                 r['combined_signal'], r['volume_ratio'], r['adx'],
                 r['momentum_5d'], r['market_cap_category'].split('(')[0].strip()]
                for r in combined]
        write_enhanced_signal_sheet(wb, f"{prefix} Combined", rows, columns_combo_enh, widths_combo_enh)

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


def main():
    print("=" * 60)
    print("  NASDAQ STOCK SCREENER - ENHANCED EDITION")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    # 1. Get all NASDAQ tickers
    all_tickers = get_nasdaq_tickers()

    # 2. Download & analyze all NASDAQ
    all_results = download_and_analyze(all_tickers, "All NASDAQ")

    # 3. Sector-specific scans
    already_analyzed = {r['ticker'] for r in all_results}

    re_tickers_missing = [t for t in RENEWABLE_ENERGY_TICKERS if t not in already_analyzed]
    re_extra = download_and_analyze(re_tickers_missing, "Renewable (extra)") if re_tickers_missing else []
    renewable_results = [r for r in all_results if r['ticker'] in RENEWABLE_ENERGY_TICKERS] + re_extra

    nuc_tickers_missing = [t for t in NUCLEAR_ENERGY_TICKERS if t not in already_analyzed]
    nuc_extra = download_and_analyze(nuc_tickers_missing, "Nuclear (extra)") if nuc_tickers_missing else []
    nuclear_results = [r for r in all_results if r['ticker'] in NUCLEAR_ENERGY_TICKERS] + nuc_extra

    ai_tickers_missing = [t for t in AI_TICKERS if t not in already_analyzed]
    ai_extra = download_and_analyze(ai_tickers_missing, "AI (extra)") if ai_tickers_missing else []
    ai_results = [r for r in all_results if r['ticker'] in AI_TICKERS] + ai_extra

    # 4. Build Excel output
    output_path = sys.argv[1] if len(sys.argv) > 1 else "nasdaq_screener_enhanced.xlsx"
    build_excel(all_results, renewable_results, nuclear_results, ai_results, output_path)

    # Print summary
    print("\n" + "=" * 60)
    print("  SUMMARY")
    print("=" * 60)
    for label, results in [("All NASDAQ", all_results), ("Renewable", renewable_results),
                           ("Nuclear", nuclear_results), ("AI", ai_results)]:
        high_score = len(filter_signals(results, 'high_score'))
        ma1 = len(filter_signals(results, 'ma50_150'))
        ma2 = len(filter_signals(results, 'ma20_50'))
        rsi = len(filter_signals(results, 'rsi_div'))
        combo = len(filter_signals(results, 'combined'))
        print(f"  {label}: {len(results)} stocks | High Score: {high_score} | "
              f"MA50×150: {ma1} | MA20×50: {ma2} | RSI Div: {rsi} | Combined: {combo}")
    print("=" * 60)
    print("\n💡 TIP: Open the Excel and check the 'TOP PICKS' sheet first!")


if __name__ == "__main__":
    main()
